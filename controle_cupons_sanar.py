import sys
import os

def caminho_recurso(nome_arquivo):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, nome_arquivo)
    return nome_arquivo
from customtkinter import CTkFont
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
from customtkinter import CTkComboBox as ttk_ComboBox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv
import matplotlib.pyplot as plt
import tkinter as tk
# Importa o módulo necessário para exibir imagens
from PIL import Image

ctk.set_appearance_mode("System")  # ou "Dark"
ctk.set_default_color_theme("blue")

# Lista de lojas cadastradas
LOJAS_CADASTRADAS = []
# Arquivo CSV para lojas
LOJAS_CSV = "lojas_cadastradas.csv"
# Arquivo CSV para indústrias
INDUSTRIAS_CSV = "industrias_cadastradas.csv"
# Lista de indústrias cadastradas
INDUSTRIAS_CADASTRADAS = []


def carregar_lojas():
    if os.path.exists(LOJAS_CSV):
        with open(caminho_recurso(LOJAS_CSV), newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if row:  # evita linhas vazias
                    LOJAS_CADASTRADAS.append(row[0])
        LOJAS_CADASTRADAS.sort()
    else:
        with open(caminho_recurso(LOJAS_CSV), mode='w', newline='', encoding='utf-8') as csvfile:
            pass  # cria o arquivo se não existir


def carregar_industrias():
    if os.path.exists(INDUSTRIAS_CSV):
        with open(caminho_recurso(INDUSTRIAS_CSV), newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if row:  # evita linhas vazias
                    INDUSTRIAS_CADASTRADAS.append(row[0])
        INDUSTRIAS_CADASTRADAS.sort()
    else:
        with open(caminho_recurso(INDUSTRIAS_CSV), mode='w', newline='', encoding='utf-8') as csvfile:
            pass  # cria o arquivo se não existir


 # Nome do arquivo Excel
EXCEL_FILE = "cupons_sanar.xlsx"

# Dashboards e resetar planilha


def abrir_dashboard():
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from collections import Counter
    if not os.path.exists(EXCEL_FILE):
        CTkMessagebox(
            title="Erro", message="Arquivo de cupons não encontrado.", icon="cancel")
        return

    wb = load_workbook(caminho_recurso(EXCEL_FILE))
    ws = wb["Cupons"]

    # Carregar todos os cupons
    cupons = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        loja, industria, valor, qtd_cupons, datahora = row
        cupons.append({
            "Loja": loja,
            "Indústria": industria,
            "Valor": valor,
            "Quantidade": qtd_cupons,
            "DataHora": datahora
        })

    # Cria nova janela para dashboard
    janela_dashboard = tk.Toplevel(janela)
    janela_dashboard.title("Dashboard de Cupons")
    janela_dashboard.update_idletasks()
    largura_janela = 900
    altura_janela = 600
    largura_tela = janela_dashboard.winfo_screenwidth()
    altura_tela = janela_dashboard.winfo_screenheight()
    x = (largura_tela // 2) - (largura_janela // 2)
    y = (altura_tela // 2) - (altura_janela // 2)
    janela_dashboard.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")

    # ==================== TOTALIZADORES E EXPORTAÇÃO PDF ====================
    # Criar frame principal para widgets acima dos gráficos
    frame_dashboard = ctk.CTkFrame(janela_dashboard)
    frame_dashboard.pack(fill="both", expand=True, padx=10, pady=10)
    # Frame para totalizadores e exportação PDF
    scrollable_frame = frame_dashboard

    # Para totalizadores, converte cupons em DataFrame
    import pandas as pd
    df = pd.DataFrame([{
        "Loja": cupom["Loja"],
        "Indústria": cupom["Indústria"],
        "Valor": cupom["Valor"],
        "Cupons": cupom["Quantidade"],
        "DataHora": cupom["DataHora"]
    } for cupom in cupons])

    # Totalizadores
    total_compras = df['Valor'].sum()
    total_cupons = df['Cupons'].sum()

    total_frame = ctk.CTkFrame(scrollable_frame)
    total_frame.pack(pady=10, fill="x")

    total_label = ctk.CTkLabel(total_frame, text=f"Total em Compras: R$ {total_compras:,.2f}", font=(
        'Arial', 14, 'bold'), anchor="w")
    total_label.pack(pady=5, fill="x")

    cupons_label = ctk.CTkLabel(total_frame, text=f"Total de Cupons: {total_cupons}", font=(
        'Arial', 14, 'bold'), anchor="w")
    cupons_label.pack(pady=5, fill="x")

    # Botão de exportar para PDF
    def exportar_dashboard_pdf():
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from datetime import datetime

        data = [['Loja', 'Indústria', 'Valor (R$)', 'Cupons']]
        grouped = df.groupby(['Loja', 'Indústria']).agg(
            {'Valor': 'sum', 'Cupons': 'sum'}).reset_index()
        for _, row in grouped.iterrows():
            data.append([row['Loja'], row['Indústria'],
                        f"{row['Valor']:.2f}", int(row['Cupons'])])

        data.append(
            ['', '', f"Total: R$ {total_compras:.2f}", f"{total_cupons} cupons"])

        doc = SimpleDocTemplate("dashboard_cupons.pdf", pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        elements.append(
            Paragraph("Dashboard de Cupons - Relatório", styles['Title']))
        elements.append(Spacer(1, 12))

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)

    exportar_btn = ctk.CTkButton(
        scrollable_frame, text="Exportar PDF", command=exportar_dashboard_pdf, anchor="w")
    exportar_btn.pack(pady=10, fill="x")

    # Conte os cupons por loja e indústria
    lojas = [cupom["Loja"] for cupom in cupons]
    industrias = [cupom["Indústria"] for cupom in cupons]
    cupons_por_loja = Counter(lojas)
    cupons_por_industria = Counter(industrias)

    # Criação da figura com subplots
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
    fig.suptitle("Análise de Cupons", fontsize=16, weight='bold')

    # Ajuste para evitar corte nos nomes das lojas no eixo X
    fig.subplots_adjust(bottom=0.3)

    # Gráfico de cupons por loja
    ax1.bar(cupons_por_loja.keys(), cupons_por_loja.values(),
            color="#5dade2", edgecolor="black")
    ax1.set_title("Cupons por Loja", fontsize=12)
    ax1.set_ylabel("Quantidade de Cupons")
    # Ajuste dos ticks para nomes de lojas longos ou muitos
    ax1.set_xticks(range(len(cupons_por_loja)))
    ax1.set_xticklabels(list(cupons_por_loja.keys()),
                        rotation=45, ha='right', fontsize=8)
    ax1.tick_params(axis='x', labelsize=8)
    ax1.grid(axis='y', linestyle='--', alpha=0.7)

    # Gráfico de cupons por indústria
    ax2.bar(cupons_por_industria.keys(), cupons_por_industria.values(),
            color="#f1948a", edgecolor="black")
    ax2.set_title("Cupons por Indústria", fontsize=12)
    ax2.tick_params(axis='x', rotation=45)
    ax2.grid(axis='y', linestyle='--', alpha=0.7)

    # Inserir no painel
    canvas = FigureCanvasTkAgg(fig, master=janela_dashboard)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

    # Ajuste para evitar corte dos rótulos do eixo X
    fig.tight_layout()


def resetar_planilha():
    msg = CTkMessagebox(title="Confirmar", message="Tem certeza que deseja resetar a planilha? Todos os dados serão apagados.",
                        icon="question", option_1="Sim", option_2="Não")
    resposta = msg.get() == "Sim"
    if resposta:
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
            CTkMessagebox(title="Resetado",
                          message="Planilha resetada com sucesso.", icon="info")
        else:
            CTkMessagebox(
                title="Info", message="Nenhuma planilha encontrada para deletar.", icon="info")

# Função que cria o arquivo Excel com cabeçalho, se ainda não existir


def criar_planilha_se_necessario():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cupons"
        ws.append(["Loja", "Indústria", "Valor da Compra (R$)",
                  "Quantidade de Cupons", "Data e Hora"])
        wb.save(caminho_recurso(EXCEL_FILE))

# Função chamada ao clicar no botão "Registrar"


def registrar_cupom():
    usar_valor_flag = usar_valor.get()
    loja = loja_var.get().strip()
    industria = industria_var.get().strip()

    if usar_valor_flag:
        valor_str = entry_valor.get().strip()
        if not valor_str:
            CTkMessagebox(
                title="Erro", message="Informe o valor da compra.", icon="cancel")
            return
        try:
            valor = float(valor_str)
            if valor < 0:
                raise ValueError
            cupons = int(valor // 500)
        except ValueError:
            CTkMessagebox(
                title="Erro", message="Insira um valor numérico válido.", icon="cancel")
            return
    else:
        cupons_str = entry_cupons.get().strip()
        if not cupons_str:
            CTkMessagebox(
                title="Erro", message="Informe a quantidade de cupons.", icon="cancel")
            return
        try:
            cupons = int(cupons_str)
            if cupons <= 0:
                raise ValueError
            valor = cupons * 500  # valor calculado automaticamente
        except ValueError:
            CTkMessagebox(
                title="Erro", message="Quantidade de cupons inválida.", icon="cancel")
            return

    # Validações básicas
    if not loja or not industria:
        CTkMessagebox(
            title="Erro", message="Todos os campos devem ser preenchidos.", icon="cancel")
        return

    # Cálculo dos cupons
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Adiciona ao Excel
    criar_planilha_se_necessario()
    wb = load_workbook(caminho_recurso(EXCEL_FILE))
    ws = wb.active
    ws.append([loja, industria, valor, cupons, data_hora])
    try:
        wb.save(caminho_recurso(EXCEL_FILE))
    except PermissionError:
        CTkMessagebox(title="Erro", message="Não foi possível salvar a planilha. Feche o arquivo 'cupons_sanar.xlsx' se ele estiver aberto no Excel e tente novamente.", icon="cancel")
        return

    # Gera uma aba de resumo com o total de compras e cupons por loja
    # Atualiza o resumo por loja
    resumo = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        loja_nome = row[0]
        valor_compra = float(row[2])
        cupons_gerados = int(row[3])
        if loja_nome in resumo:
            resumo[loja_nome]["valor_total"] += valor_compra
            resumo[loja_nome]["cupons_total"] += cupons_gerados
        else:
            resumo[loja_nome] = {
                "valor_total": valor_compra,
                "cupons_total": cupons_gerados
            }

    # Remove aba "Resumo" se já existir
    if "Resumo" in wb.sheetnames:
        del wb["Resumo"]

    # Cria nova aba de resumo
    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Loja", "Total em Compras (R$)", "Total de Cupons"])
    for loja_nome, dados in resumo.items():
        ws_resumo.append(
            [loja_nome, dados["valor_total"], dados["cupons_total"]])

    # Resumo por indústria
    resumo_industria = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        industria_nome = row[1]
        valor_compra = float(row[2])
        cupons_gerados = int(row[3])
        if industria_nome in resumo_industria:
            resumo_industria[industria_nome]["valor_total"] += valor_compra
            resumo_industria[industria_nome]["cupons_total"] += cupons_gerados
        else:
            resumo_industria[industria_nome] = {
                "valor_total": valor_compra,
                "cupons_total": cupons_gerados
            }

    if "Resumo_Indústrias" in wb.sheetnames:
        del wb["Resumo_Indústrias"]
    ws_industrias = wb.create_sheet("Resumo_Indústrias")
    ws_industrias.append(
        ["Indústria", "Total em Compras (R$)", "Total de Cupons"])
    for industria, dados in resumo_industria.items():
        ws_industrias.append(
            [industria, dados["valor_total"], dados["cupons_total"]])

    # Totais gerais
    total_compras_geral = sum(
        [float(row[2]) for row in ws.iter_rows(min_row=2, values_only=True)])
    total_cupons_geral = sum(
        [int(row[3]) for row in ws.iter_rows(min_row=2, values_only=True)])

    if "Totais_Gerais" in wb.sheetnames:
        del wb["Totais_Gerais"]
    ws_totais = wb.create_sheet("Totais_Gerais")
    ws_totais.append(["Total Geral de Compras (R$)", total_compras_geral])
    ws_totais.append(["Total Geral de Cupons", total_cupons_geral])

    try:
        wb.save(caminho_recurso(EXCEL_FILE))
    except PermissionError:
        CTkMessagebox(title="Erro", message="Não foi possível salvar a planilha. Feche o arquivo 'cupons_sanar.xlsx' se ele estiver aberto no Excel e tente novamente.", icon="cancel")
        return

    # Limpa os campos
    entry_valor.delete(0, tk.END)
    entry_cupons.delete(0, tk.END)
    janela.update_idletasks()
    CTkMessagebox(
        title="Sucesso", message=f"{cupons} cupom(ns) registrado(s) com sucesso!", icon="info")
    # Atualiza consulta se a janela estiver aberta
    for toplevel in janela.winfo_children():
        if isinstance(toplevel, ctk.CTkToplevel) and toplevel.title() == "Consulta de Cupons":
            for widget in toplevel.winfo_children():
                if hasattr(widget, "aplicar_filtro"):
                    widget.aplicar_filtro()


def excluir_cupom():
    loja = loja_var.get().strip()
    industria = industria_var.get().strip()
    if not loja or not industria:
        CTkMessagebox(
            title="Erro", message="Selecione a loja e a indústria para excluir os cupons.", icon="cancel")
        return

    if not os.path.exists(EXCEL_FILE):
        CTkMessagebox(
            title="Erro", message="Arquivo de cupons não encontrado.", icon="cancel")
        return

    wb = load_workbook(caminho_recurso(EXCEL_FILE))
    ws = wb.active
    linhas_para_remover = []

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == loja and row[1].value == industria:
            linhas_para_remover.append(i)

    if not linhas_para_remover:
        CTkMessagebox(
            title="Info", message="Nenhum cupom encontrado para os critérios selecionados.", icon="info")
        return

    for i in reversed(linhas_para_remover):
        ws.delete_rows(i)

    # Remove e recria as abas de resumo para manter os dados atualizados
    if "Resumo" in wb.sheetnames:
        del wb["Resumo"]
    if "Resumo_Indústrias" in wb.sheetnames:
        del wb["Resumo_Indústrias"]
    if "Totais_Gerais" in wb.sheetnames:
        del wb["Totais_Gerais"]

    # Recria as abas de resumo (reaproveitando lógica já existente)
    resumo = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        loja_nome = row[0]
        valor_compra = float(row[2])
        cupons_gerados = int(row[3])
        if loja_nome in resumo:
            resumo[loja_nome]["valor_total"] += valor_compra
            resumo[loja_nome]["cupons_total"] += cupons_gerados
        else:
            resumo[loja_nome] = {
                "valor_total": valor_compra,
                "cupons_total": cupons_gerados
            }

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Loja", "Total em Compras (R$)", "Total de Cupons"])
    for loja_nome, dados in resumo.items():
        ws_resumo.append(
            [loja_nome, dados["valor_total"], dados["cupons_total"]])

    resumo_industria = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        industria_nome = row[1]
        valor_compra = float(row[2])
        cupons_gerados = int(row[3])
        if industria_nome in resumo_industria:
            resumo_industria[industria_nome]["valor_total"] += valor_compra
            resumo_industria[industria_nome]["cupons_total"] += cupons_gerados
        else:
            resumo_industria[industria_nome] = {
                "valor_total": valor_compra,
                "cupons_total": cupons_gerados
            }

    ws_industrias = wb.create_sheet("Resumo_Indústrias")
    ws_industrias.append(
        ["Indústria", "Total em Compras (R$)", "Total de Cupons"])
    for industria, dados in resumo_industria.items():
        ws_industrias.append(
            [industria, dados["valor_total"], dados["cupons_total"]])

    total_compras_geral = sum(
        [float(row[2]) for row in ws.iter_rows(min_row=2, values_only=True)])
    total_cupons_geral = sum(
        [int(row[3]) for row in ws.iter_rows(min_row=2, values_only=True)])

    ws_totais = wb.create_sheet("Totais_Gerais")
    ws_totais.append(["Total Geral de Compras (R$)", total_compras_geral])
    ws_totais.append(["Total Geral de Cupons", total_cupons_geral])

    try:
        wb.save(caminho_recurso(EXCEL_FILE))
    except PermissionError:
        CTkMessagebox(title="Erro", message="Não foi possível salvar a planilha. Feche o arquivo 'cupons_sanar.xlsx' se ele estiver aberto no Excel e tente novamente.", icon="cancel")
        return

    CTkMessagebox(title="Sucesso",
                  message="Cupons excluídos com sucesso.", icon="info")
    # Atualiza consulta se a janela estiver aberta
    for toplevel in janela.winfo_children():
        if isinstance(toplevel, ctk.CTkToplevel) and toplevel.title() == "Consulta de Cupons":
            for widget in toplevel.winfo_children():
                if hasattr(widget, "aplicar_filtro"):
                    widget.aplicar_filtro()
                    toplevel.update_idletasks()


carregar_lojas()
carregar_industrias()

janela = ctk.CTk()
# Função de encerramento limpa


def ao_fechar_janela():
    try:
        janela.destroy()
    except:
        pass
    janela.quit()


janela.protocol("WM_DELETE_WINDOW", ao_fechar_janela)
janela.title("Controle de Cupons - Rede Sanar")
# Centralizar a janela principal e definir tamanho dinâmico baseado na resolução
largura_tela = janela.winfo_screenwidth()
altura_tela = janela.winfo_screenheight()
largura_janela = int(largura_tela * 0.5)
altura_janela = int(altura_tela * 0.6)
x_pos = (largura_tela - largura_janela) // 2
y_pos = (altura_tela - altura_janela) // 2
janela.geometry(f"{largura_janela}x{altura_janela}+{x_pos}+{y_pos}")
janela.resizable(True, True)

# Frame principal para responsividade
main_frame = ctk.CTkFrame(janela)
main_frame.pack(fill="both", expand=True, padx=10, pady=10)

# Rótulos e campos
ctk.CTkLabel(main_frame, text="Loja:").pack(pady=(10, 0))
loja_var = ctk.StringVar(
    value=LOJAS_CADASTRADAS[0] if LOJAS_CADASTRADAS else "")
option_loja = ctk.CTkOptionMenu(
    main_frame, variable=loja_var, values=LOJAS_CADASTRADAS)
option_loja.pack()

ctk.CTkLabel(main_frame, text="Indústria:").pack(pady=(10, 0))
industria_var = ctk.StringVar(
    value=INDUSTRIAS_CADASTRADAS[0] if INDUSTRIAS_CADASTRADAS else "")
option_industria = ctk.CTkOptionMenu(
    main_frame, variable=industria_var, values=INDUSTRIAS_CADASTRADAS if INDUSTRIAS_CADASTRADAS else [""])
option_industria.pack()

ctk.CTkLabel(main_frame, text="Valor da Compra (R$):").pack(pady=(10, 0))
entry_valor = ctk.CTkEntry(main_frame, width=200)
entry_valor.pack()
entry_valor.bind("<Return>", lambda event: registrar_cupom())

usar_valor = ctk.BooleanVar(value=True)
ctk.CTkCheckBox(main_frame, text="Usar Valor", variable=usar_valor).pack()

ctk.CTkLabel(main_frame, text="Ou Quantidade de Cupons:").pack(pady=(10, 0))
entry_cupons = ctk.CTkEntry(main_frame, width=200)
entry_cupons.pack()
entry_cupons.bind("<Return>", lambda event: registrar_cupom())

# Botão de registrar
ctk.CTkButton(main_frame, text="Registrar Cupom",
              command=registrar_cupom, width=150, height=40).pack(pady=20)


def abrir_tela_consulta():
    if not os.path.exists(EXCEL_FILE):
        CTkMessagebox(
            title="Erro", message="Arquivo de cupons não encontrado.", icon="cancel")
        return

    consulta = ctk.CTkToplevel(janela)
    consulta.title("Consulta de Cupons")
    largura_tela = consulta.winfo_screenwidth()
    altura_tela = consulta.winfo_screenheight()
    largura_consulta = int(largura_tela * 0.7)
    altura_consulta = int(altura_tela * 0.7)
    pos_x = (largura_tela - largura_consulta) // 2
    pos_y = (altura_tela - altura_consulta) // 2
    consulta.geometry(f"{largura_consulta}x{altura_consulta}+{pos_x}+{pos_y}")
    consulta.resizable(True, True)

    # Frame principal para responsividade
    consulta_main_frame = ctk.CTkFrame(consulta)
    consulta_main_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # Frame para filtros
    frame_filtros = ctk.CTkFrame(consulta_main_frame)
    frame_filtros.pack(fill="x", pady=10)

    wb = load_workbook(caminho_recurso(EXCEL_FILE))
    ws = wb["Cupons"]

    lojas_disponiveis = sorted(
        set(row[0] for row in ws.iter_rows(min_row=2, values_only=True)))
    industrias_disponiveis = sorted(
        set(row[1] for row in ws.iter_rows(min_row=2, values_only=True)))

    ctk.CTkLabel(frame_filtros, text="Filtrar por Loja:", anchor="w").grid(
        row=0, column=0, padx=5, sticky="w")
    filtro_loja_var = ctk.StringVar(value="")
    filtro_loja_menu = ctk.CTkOptionMenu(
        frame_filtros, variable=filtro_loja_var, values=[""] + lojas_disponiveis)
    filtro_loja_menu.grid(row=0, column=1, padx=5, sticky="ew")

    ctk.CTkLabel(frame_filtros, text="Filtrar por Indústria:", anchor="w").grid(
        row=0, column=2, padx=5, sticky="w")
    filtro_industria_var = ctk.StringVar(value="")
    filtro_industria_menu = ctk.CTkOptionMenu(
        frame_filtros, variable=filtro_industria_var, values=[""] + industrias_disponiveis)
    filtro_industria_menu.grid(row=0, column=3, padx=5, sticky="ew")
    frame_filtros.grid_columnconfigure(1, weight=1)
    frame_filtros.grid_columnconfigure(3, weight=1)

    # Frame para resultados (Treeview com rolagem)
    consulta_frame = ctk.CTkFrame(consulta_main_frame)
    consulta_frame.pack(fill="both", expand=True)
    consulta_frame.grid_rowconfigure(0, weight=1)
    consulta_frame.grid_columnconfigure(0, weight=1)

    from tkinter import ttk
    # Treeview com rolagem
    tree = ttk.Treeview(consulta_frame, columns=(
        "Loja", "Indústria", "Valor", "Cupons", "DataHora"), show="headings")
    for coluna in ("Loja", "Indústria", "Valor", "Cupons", "DataHora"):
        tree.heading(coluna, text={
            "Loja": "Loja",
            "Indústria": "Indústria",
            "Valor": "Valor da Compra (R$)",
            "Cupons": "Quantidade de Cupons",
            "DataHora": "Data e Hora"
        }[coluna])
        tree.column(coluna, anchor="w", width=150, stretch=True)

    # Scrollbar vertical
    scrollbar_y = ttk.Scrollbar(
        consulta_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar_y.set)
    tree.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    scrollbar_y.grid(row=0, column=1, sticky="ns")
    consulta_frame.grid_rowconfigure(0, weight=1)
    consulta_frame.grid_columnconfigure(0, weight=1)

    # Frame para totais
    total_frame = ctk.CTkFrame(consulta_main_frame)
    total_frame.pack(pady=10, fill="x")

    total_valor_label = ctk.CTkLabel(
        total_frame, text="Total em Compras (R$): 0.00", anchor="w")
    total_valor_label.pack(fill="x")

    total_cupons_label = ctk.CTkLabel(
        total_frame, text="Total de Cupons: 0", anchor="w")
    total_cupons_label.pack(fill="x")

    def aplicar_filtro():
        for item in tree.get_children():
            tree.delete(item)
        filtro_loja = filtro_loja_var.get().lower()
        filtro_industria = filtro_industria_var.get().lower()
        total_valor = 0
        total_cupons = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            loja_excel = row[0].strip().lower()
            industria_excel = row[1].strip().lower()
            if (filtro_loja in loja_excel) and (filtro_industria in industria_excel):
                tree.insert("", tk.END, values=row)
                total_valor += float(row[2])
                total_cupons += int(row[3])
        total_valor_label.configure(
            text=f"Total em Compras (R$): {total_valor:.2f}")
        total_cupons_label.configure(text=f"Total de Cupons: {total_cupons}")
        # Ajuste automático das colunas
        for col in ("Loja", "Indústria", "Valor", "Cupons", "DataHora"):
            tree.column(col, width=150, stretch=True)
            # Ajuste dinâmico de largura
            tree.update_idletasks()
            maxwidth = 150
            for iid in tree.get_children():
                text = str(tree.set(iid, col))
                w = max(tree.column(col, 'width'), len(text)*8)
                if w > maxwidth:
                    maxwidth = w
            tree.column(col, width=maxwidth)
    consulta.aplicar_filtro = aplicar_filtro

    def exportar_filtro_pdf():
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        import tempfile
        import os
        dados_exportar = []
        headers = ["Loja", "Indústria",
                   "Valor da Compra (R$)", "Quantidade de Cupons", "Data e Hora"]
        for item in tree.get_children():
            dados_exportar.append(tree.item(item)["values"])
        if not dados_exportar:
            CTkMessagebox(
                title="Aviso", message="Nenhum dado para exportar.", icon="warning")
            return
        styles = getSampleStyleSheet()
        pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        doc = SimpleDocTemplate(
            pdf_file.name, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=50)
        elements = [Paragraph("Relatório de Cupons Filtrados",
                              styles['Heading1']), Spacer(1, 12)]
        # Inserir logo centralizada no topo, se existir
        logo_path = caminho_recurso("logo_sanar.png")
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=100, height=50)
            logo.hAlign = 'CENTER'
            elements.insert(0, Spacer(1, 10))
            elements.insert(0, logo)
        # Formatar dados com Parágrafos para quebra de linha automática
        tabela_formatada = [headers]
        for linha in dados_exportar:
            linha_formatada = [
                Paragraph(str(c), styles['Normal']) for c in linha]
            tabela_formatada.append(linha_formatada)
        col_widths = [130, 130, 100, 120, 110]
        table = Table(tabela_formatada, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(table)
        total_valor = sum(float(linha[2]) for linha in dados_exportar)
        total_cupons = sum(int(linha[3]) for linha in dados_exportar)
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(f"Total em Compras (R$): {total_valor:.2f}", styles["Heading3"]))
        elements.append(
            Paragraph(f"Total de Cupons: {total_cupons}", styles["Heading3"]))
        try:
            doc.build(elements)
            os.system(f"open {pdf_file.name}")
        except Exception as e:
            CTkMessagebox(
                title="Erro", message=f"Erro ao gerar PDF: {str(e)}", icon="cancel")

    # Botões com tamanho adaptável
    btn_frame = ctk.CTkFrame(consulta_main_frame)
    btn_frame.pack(pady=(0, 10), fill="x")
    ctk.CTkButton(btn_frame, text="Aplicar Filtro", command=aplicar_filtro,
                  width=150, height=40).pack(side="left", padx=10, pady=5)
    ctk.CTkButton(btn_frame, text="Exportar Filtro para PDF", command=exportar_filtro_pdf,
                  width=150, height=40).pack(side="left", padx=10, pady=5)

    def excluir_cupom_selecionado():
        item = tree.selection()
        if not item:
            CTkMessagebox(
                title="Aviso", message="Selecione um cupom para excluir.", icon="warning")
            return
        msg = CTkMessagebox(
            title="Confirmar",
            message="Deseja realmente excluir este cupom?",
            icon="question",
            option_1="Sim",
            option_2="Não"
        )
        confirm = msg.get() == "Sim"
        if not confirm:
            return
        dados = tree.item(item[0])["values"]
        if not os.path.exists(EXCEL_FILE):
            CTkMessagebox(
                title="Erro", message="Arquivo de cupons não encontrado.", icon="cancel")
            return
        wb = load_workbook(caminho_recurso(EXCEL_FILE))
        ws = wb["Cupons"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            loja_cell = str(row[0].value).strip()
            industria_cell = str(row[1].value).strip()
            valor_cell = float(row[2].value)
            cupons_cell = int(row[3].value)
            data_cell = str(row[4].value).strip()
            if (
                loja_cell == str(dados[0]).strip()
                and industria_cell == str(dados[1]).strip()
                and abs(valor_cell - float(dados[2])) < 0.01
                and cupons_cell == int(dados[3])
                and data_cell == str(dados[4]).strip()
            ):
                ws.delete_rows(i)
                wb.save(caminho_recurso(EXCEL_FILE))
                wb.close()
                wb = load_workbook(EXCEL_FILE)
                ws = wb["Cupons"]
                CTkMessagebox(
                    title="Sucesso", message="Cupom excluído com sucesso!", icon="info")
                aplicar_filtro()
                consulta.update_idletasks()
                return
    ctk.CTkButton(btn_frame, text="Excluir Cupom Selecionado",
                  command=excluir_cupom_selecionado, width=150, height=40).pack(side="left", padx=10, pady=5)


def abrir_tela_cadastro():
    cadastro = ctk.CTkToplevel(janela)
    cadastro.title("Cadastro de Lojas e Indústrias")
    cadastro.update_idletasks()
    largura_janela = 900
    altura_janela = 600
    largura_tela = cadastro.winfo_screenwidth()
    altura_tela = cadastro.winfo_screenheight()
    x = (largura_tela // 2) - (largura_janela // 2)
    y = (altura_tela // 2) - (altura_janela // 2)
    cadastro.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")

    frame = ctk.CTkFrame(cadastro)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    frame_loja = ctk.CTkFrame(frame)
    frame_loja.pack(pady=10, fill="x", expand=True)
    ctk.CTkLabel(frame_loja, text="Nova Loja:", anchor="w").pack(fill="x")
    nova_loja_entry = ctk.CTkEntry(frame_loja)
    nova_loja_entry.pack(fill="x")

    def salvar_loja():
        nova_loja = nova_loja_entry.get().strip()
        if nova_loja and nova_loja not in LOJAS_CADASTRADAS:
            with open(caminho_recurso(LOJAS_CSV), "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([nova_loja])
            LOJAS_CADASTRADAS.append(nova_loja)
            option_loja.configure(values=sorted(LOJAS_CADASTRADAS))
            CTkMessagebox(title="Sucesso",
                          message="Loja cadastrada com sucesso!", icon="info")
            nova_loja_entry.delete(0, tk.END)
    ctk.CTkButton(frame_loja, text="Salvar Loja",
                  command=salvar_loja, width=150, height=40, anchor="w").pack(pady=5, fill="x")

    frame_ind = ctk.CTkFrame(frame)
    frame_ind.pack(pady=10, fill="x", expand=True)
    ctk.CTkLabel(frame_ind, text="Nova Indústria:", anchor="w").pack(fill="x")
    nova_ind_entry = ctk.CTkEntry(frame_ind)
    nova_ind_entry.pack(fill="x")

    def salvar_industria():
        nova_ind = nova_ind_entry.get().strip()
        if nova_ind and nova_ind not in INDUSTRIAS_CADASTRADAS:
            with open(caminho_recurso(INDUSTRIAS_CSV), "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([nova_ind])
            INDUSTRIAS_CADASTRADAS.append(nova_ind)
            option_industria.configure(values=sorted(INDUSTRIAS_CADASTRADAS))
            CTkMessagebox(
                title="Sucesso", message="Indústria cadastrada com sucesso!", icon="info")
            nova_ind_entry.delete(0, tk.END)
    ctk.CTkButton(frame_ind, text="Salvar Indústria",
                  command=salvar_industria, width=150, height=40, anchor="w").pack(pady=5, fill="x")


# Botão para abrir a tela de consulta de cupons
ctk.CTkButton(main_frame, text="Consultar Cupons",
              command=abrir_tela_consulta, width=150, height=40).pack(pady=5)

# Novo frame lateral para botões menos relevantes
side_frame = ctk.CTkFrame(janela, fg_color="transparent")
side_frame.pack(side="left", fill="y", padx=(0, 5), pady=10, anchor="n")

button_font = CTkFont(size=12)

ctk.CTkButton(
    side_frame,
    text="Dashboards",
    command=abrir_dashboard,
    width=120,
    height=32,
    corner_radius=5,
    fg_color="gray20",
    hover_color="gray30",
    font=button_font
).pack(pady=(5, 2), padx=10, fill="x")
ctk.CTkButton(
    side_frame,
    text="Resetar Planilha",
    command=resetar_planilha,
    width=120,
    height=32,
    corner_radius=5,
    fg_color="gray20",
    hover_color="gray30",
    font=button_font
).pack(pady=2, padx=10, fill="x")
ctk.CTkButton(
    side_frame,
    text="Cadastrar Lojas/Indústrias",
    command=abrir_tela_cadastro,
    width=120,
    height=32,
    corner_radius=5,
    fg_color="gray20",
    hover_color="gray30",
    font=button_font
).pack(pady=2, padx=10, fill="x")


# Inicia a interface
janela.mainloop()
